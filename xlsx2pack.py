#!/usr/bin/env python3
# แปลง checklist.xlsx (แม่แบบมาตรฐาน) -> appdata.json ของ pack
# ใช้เอง:  python3 xlsx2pack.py packs/<proj>/<ver>/checklist.xlsx
# หรือปล่อยให้ build.py เรียกอัตโนมัติเมื่อ checklist.xlsx ใหม่กว่า appdata.json
#
# คอลัมน์ (แถวแรกเป็นหัวตาราง — ห้ามเปลี่ยนชื่อ):
#  Module | ModuleTitle | ModuleTH | ManualRef | Part | PartTH | Section | GroupNo | GroupSubject
#  | ItemID | Item_EN | Item_TH | RegRef | RefKey | Evidence | HowToVerify | Dual(Y/N)
# - แถวข้อมูล 1 แถว = รายการตรวจ 1 ข้อ · ช่อง Module/Part/Section/Group เว้นว่างได้ = ใช้ค่าเดิมจากแถวก่อนหน้า
# - ItemID เว้นว่าง = ออกเลขอัตโนมัติ <Module>-0001 ต่อเนื่อง
import sys, os, json
import openpyxl

COLS = ['Module','ModuleTitle','ModuleTH','ManualRef','Part','PartTH','Section',
        'GroupNo','GroupSubject','ItemID','Item_EN','Item_TH','RegRef','RefKey',
        'Evidence','HowToVerify','SubChecklist','Dual']

def convert(xlsx_path, out_path=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['checklist'] if 'checklist' in wb.sheetnames else wb[wb.sheetnames[0]]
    head = [str(c.value).strip() if c.value else '' for c in ws[1]]
    idx = {}
    for c in COLS:
        if c in head: idx[c] = head.index(c)
    missing = [c for c in ('Module','Item_EN') if c not in idx]
    if missing: raise SystemExit(f'{xlsx_path}: ไม่พบคอลัมน์ {missing} — ใช้หัวตารางตามแม่แบบ')
    def get(row, col):
        i = idx.get(col)
        if i is None or i >= len(row): return ''
        v = row[i].value
        return str(v).strip() if v is not None else ''

    modules, mod_by_code, counters = [], {}, {}
    cur = dict.fromkeys(COLS, '')
    for row in ws.iter_rows(min_row=2):
        vals = {c: get(row, c) for c in COLS}
        if not any(vals.values()): continue
        for c in ('Module','ModuleTitle','ModuleTH','ManualRef','Part','PartTH','Section','GroupNo','GroupSubject'):
            if vals[c]: cur[c] = vals[c]
            vals[c] = cur[c]
        if not vals['Item_EN']: continue
        mc = vals['Module']
        m = mod_by_code.get(mc)
        if not m:
            m = {'c': mc, 't': vals['ModuleTitle'] or mc, 'th': vals['ModuleTH'],
                 'doc': '', 'man': vals['ManualRef'], 'parts': []}
            mod_by_code[mc] = m; modules.append(m); counters[mc] = 0
        pt = vals['Part'] or 'Part 1'
        p = next((x for x in m['parts'] if x['t'] == pt), None)
        if not p:
            p = {'t': pt, 'th': vals['PartTH'], 'secs': []}
            m['parts'].append(p)
        stt = vals['Section']
        sec = next((x for x in p['secs'] if x['t'] == stt), None)
        if not sec:
            sec = {'t': stt, 'groups': []}
            p['secs'].append(sec)
        gn, gs = vals['GroupNo'] or str(len(sec['groups'])+1), vals['GroupSubject']
        g = next((x for x in sec['groups'] if x['n'] == gn and x['subj'] == gs), None)
        if not g:
            g = {'n': gn, 'subj': gs, 'items': []}
            sec['groups'].append(g)
        counters[mc] += 1
        iid = vals['ItemID'] or f'{mc}-{counters[mc]:04d}'
        item = {'id': iid, 'd': vals['Item_EN'], 'rr': vals['RegRef'], 'rk': vals['RefKey'],
                'ref': vals['Evidence'], 'prev': '', 'hd': 0, 'how': vals['HowToVerify']}
        if vals['Item_TH']: item['th'] = vals['Item_TH']
        if vals['Dual'].upper().startswith('Y'): item['dual'] = 1
        sub=[x.strip() for x in vals['SubChecklist'].replace('|','\n').split('\n') if x.strip()]
        if sub: item['chk']=sub
        g['items'].append(item)

    data = {'modules': modules, 'catalog': []}
    out = out_path or os.path.join(os.path.dirname(xlsx_path), 'appdata.json')
    json.dump(data, open(out, 'w', encoding='utf-8'), ensure_ascii=False, separators=(',', ':'))
    n = sum(counters.values())
    print(f'xlsx2pack: {xlsx_path} -> {out} ({n} items, {len(modules)} modules)')
    return out

if __name__ == '__main__':
    if len(sys.argv) < 2: raise SystemExit('usage: xlsx2pack.py <checklist.xlsx>')
    convert(sys.argv[1])
